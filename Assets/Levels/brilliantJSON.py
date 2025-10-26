import os
from openpyxl import load_workbook
import json

#Get Folder of Python file
folder_path = os.path.dirname(os.path.abspath(__file__))

#Get first .xlsx file in folder
xlsx_files = [f for f in os.listdir(folder_path) if f.endswith("xlsx")]

if not xlsx_files:
    raise FileNotFoundError("No .xlsx files found in this directory!")

#Pick the first Excel file (alphabeticaly)
first_file = os.path.join(folder_path, xlsx_files[0])
print(first_file)

#open workbook
wb = load_workbook(first_file, data_only=False)
sheet = wb.active

levels = []
levelsBoardSize = []

for i in range(len(wb.sheetnames)):
 
    cellData = {"cells": []}
    max_row = 20
    max_column = 20

    sheet = wb[wb.sheetnames[i]]
    for row in (sheet.iter_rows()):

        #Keeping max 20 row
        if row[0].row > max_row:
            break

        for cell in row:

            # Skip columns after column 20
            if cell.column > max_column:
                continue
            
            fill_color = 'None'
            if cell.fill and cell.fill.fgColor:
                fg = cell.fill.fgColor
                if fg.type == 'rgb' and fg.rgb:
                    fill_color = fg.rgb.upper()

            data = {
                "row": cell.row,
                "column": cell.column,
                "value": cell.value,
                "file_color": fill_color
                
            }

            #find boarders
            if '`' in str(cell.value):
                if cell.column > cell.row:
                    max_column = cell.column - 1
                    continue
                else:
                    max_row = cell.row - 1
                    break

            if cell.value != None or fill_color != '00000000':
                if fill_color == "None":
                    raise FileNotFoundError(f"Theme Color ignored at row {cell.row}, Column {cell.column}")

                cellData["cells"].append(data)

    
    board_size = {"boardSize": [max_row, max_column]}
    levelsBoardSize.append(board_size)
    
    levels.append(cellData)
    

for i in range(len(wb.sheetnames)):
    output_file = os.path.join(folder_path, f"brilliantLevel{i + 1}.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump({"level": i + 1, **levelsBoardSize[i], **levels[i]}, f, indent=2, ensure_ascii=False)


print(f"✅ Exported Excel data to: {output_file}")


print(cellData)
print(max_row,max_column)