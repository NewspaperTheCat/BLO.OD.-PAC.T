import os
from openpyxl import load_workbook
import json

#Get Folder of Python file
folder_path = os.path.dirname(os.path.abspath(__file__))

#Get first .xlsx file in folder
xlsx_files = [f for f in os.listdir(folder_path) if f.endswith("xlsx")]

if not xlsx_files:
    raise FileNotFoundError("No .xlsx files found in this directory!")

#Pick the first Excel file (alphabeticaly)
first_file = os.path.join(folder_path, xlsx_files[0])
print(first_file)

#open workbook
wb = load_workbook(first_file, data_only=False)
sheet = wb.active

levels = []
levelsBoardSize = []
levelsDescription = []
levelsRequirements = []

def wordAndColorSet(cel, col, cr):

    this_cell = []

    if cel.value != None:
        this_cell.append(cel.value)
    else:
        this_cell.append("Null")

    if col != '00000000' and col != "None" :
        this_cell.append(col)
    else:
        this_cell.append("Null")

    cr.append(this_cell)

for i in range(len(wb.sheetnames)):
 
    cellData = {"cells": []}
    max_row = 20
    max_column = 20

    sheet = wb[wb.sheetnames[i]]
    for row in (sheet.iter_rows()):

        #Keeping max 20 row
        if row[0].row > max_row:
            break

        for cell in row:

            # Skip columns after column 20
            if cell.column > max_column:
                continue
            
            fill_color = 'None'
            if cell.fill and cell.fill.fgColor:
                fg = cell.fill.fgColor
                if fg.type == 'rgb' and fg.rgb:
                    fill_color = fg.rgb.upper()
                    if fill_color == "00000000" or fill_color == "FF000000":
                        fill_color = "FFFFFFFF"

            data = {
                "row": cell.row,
                "column": cell.column,
                "value": cell.value,
                "file_color": fill_color
                
            }

            #find boarders
            if '`' in str(cell.value):
                if cell.column > cell.row:
                    max_column = cell.column - 1
                    continue
                else:
                    max_row = cell.row - 1
                    descriptionCell = sheet.cell(cell.row + 1, 1)
                    if(descriptionCell.value != None):
                        levelsDescription.append({"description": descriptionCell.value})
                    else:
                        raise FileNotFoundError(f"Level Description Not Found on page {i+1}!")
                    break

            if cell.value != None or fill_color != '00000000':
                if fill_color == "None":
                    raise FileNotFoundError(f"Theme Color ignored at row {cell.row}, Column {cell.column}")

                cellData["cells"].append(data)

    board_size = {"boardSize": [max_row, max_column]}
    levelsBoardSize.append(board_size)    
    levels.append(cellData)
    
    validRequirementTypes = ["hoverover","nocell","replace", "answerkey"]

    stop = False
    requirementData = {
    "requirements": {
        "hoverover": [],
        "nocell": [],
        "replace": [],
        "answerkey": {
                "regionStart": [-1, -1],
                "regionEnd": [6, 7],
                "cells": []
            }
        }
    }
    currentRequirementData = {}

    answer_key = False
    answer_key_row = -1

    ##Read Key Requirements
    for row in sheet.iter_rows(min_row =  max_row + 3):
        for cell in row:

            # Skip columns after column 20
            if cell.column > max_column:
                continue
            
            if cell.column == 1:
                if cell._value == None:
                    stop = True
                    break
                
                requirementType = cell.value.lower().replace(" ", "")
                if requirementType not in validRequirementTypes:
                    raise FileNotFoundError(f"{requirementType} is not a Valid Requirement, on page {i+1}!")
                else:
                    currentRequirementData = requirementType
                    print("Hello?")
                    if currentRequirementData == validRequirementTypes[3]:
                        answer_key = True
                        answer_key_row = cell.row
                        stop = True

            if cell.column != 1:

                #Get Cell Background Color
                fill_color = "None"
                if cell.fill and cell.fill.fgColor:
                    fg = cell.fill.fgColor
                    if fg.type == 'rgb' and fg.rgb:
                        fill_color = fg.rgb.upper()

                #Report Error
                if cell.value == None and (fill_color == '00000000' or fill_color == "None"):
                    raise FileNotFoundError(f"{requirementType}, should not have a null value and no color on row: {cell.row}, cell: {cell.column}, on page {i+1}")

                if '`' in str(cell.value):
                    break
                if requirementType == validRequirementTypes[0]:
                    wordAndColorSet(cell, fill_color, requirementData["requirements"][currentRequirementData])
                if requirementType == validRequirementTypes[1]:
                    wordAndColorSet(cell, fill_color, requirementData["requirements"][currentRequirementData])
                if requirementType == validRequirementTypes[2]:
                    wordAndColorSet(cell, fill_color, requirementData["requirements"][currentRequirementData])
        if stop:
            break

    #AnswerKey
    if answer_key:
        hasSetSize = False
        answer_key_size = []
        answerData = {"cells": []}

        for row in sheet.iter_rows(min_row = answer_key_row):

            #setGrid
            if not hasSetSize:
                if '`' in str(row[1].value):
                    answer_key_size = [1,1, max_row, max_column]

                    requirementData["requirements"][currentRequirementData]["regionStart"] = [answer_key_size[0], answer_key_size[1]]
                    requirementData["requirements"][currentRequirementData]["regionEnd"] =  [answer_key_size[2], answer_key_size[3]]

                    hasSetSize = True
                    continue
                elif '`' in str(row[5].value):
                    if len(row[1].value) > 1 or len(row[3].value) > 1:
                        raise FileNotFoundError(f"{requirementType}, should not have a more then one letter!, on page {i+1}")
                    convertLetter1 = ord(row[1].value[0].lower()) - 96
                    convertLetter2 = ord(row[3].value[0].lower()) - 96
                    answer_key_size = [row[2].value, convertLetter1, row[4].value, convertLetter2]

                    requirementData["requirements"][currentRequirementData]["regionStart"] = [answer_key_size[0], answer_key_size[1]]
                    requirementData["requirements"][currentRequirementData]["regionEnd"] =  [answer_key_size[2], answer_key_size[3]]

                    hasSetSize = True
                    continue
                else:
                    raise FileNotFoundError(f"Answer Key Requirements formated incorrectly, on page {i+1}!")
                
        for row in sheet.iter_rows(min_row = answer_key_row + 1):
            #Keeping max 20 row
            if row[0].row > answer_key_row + answer_key_size[2] - answer_key_size[0] + 1:
                break

            for cell in row:
                  # Skip columns after column 20
                if cell.column > answer_key_size[3] - answer_key_size[1] + 1:
                    continue
                
                fill_color = 'None'
                if cell.fill and cell.fill.fgColor:
                    fg = cell.fill.fgColor
                    if fg.type == 'rgb' and fg.rgb:
                        fill_color = fg.rgb.upper()
                        if fill_color == "00000000" or fill_color == "FF000000":
                            fill_color = "00000000"

                data = {
                    "row": cell.row - answer_key_row,
                    "column": cell.column + answer_key_size[1] - 1,
                    "value": cell.value,
                    "file_color": fill_color
                    
                }
                if cell.value != None or fill_color != '00000000':
                    if fill_color == "None":
                        raise FileNotFoundError(f"Theme Color ignored at row {cell.row}, Column {cell.column}")
                    answerData["cells"].append(data)
        
            requirementData["requirements"][currentRequirementData]["cells"] = answerData["cells"]
    levelsRequirements.append(requirementData)


for i in range(len(wb.sheetnames)):
    output_file = os.path.join(folder_path, f"brilliantLevel{i + 1}.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump({"level": i + 1,**levelsDescription[i], **levelsRequirements[i], **levelsBoardSize[i], **levels[i]}, f, indent=2, ensure_ascii=False)
        ##if(i == 0):
            ##print(levels[i])

##print(f"✅ Exported Excel data to: {output_file}")
##print(max_row,max_column)
